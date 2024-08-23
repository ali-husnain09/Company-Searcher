from selenium import webdriver
from selenium.common.exceptions import (
    NoSuchElementException,
    TimeoutException,
    StaleElementReferenceException,
)
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import json
import openpyxl
import colorama
from colorama import Fore, Back
import undetected_chromedriver as uc
from address_check import address_validations
from name_check import name_validations
import os

colorama.init(autoreset=True)


class CompanySearches:
    def __init__(self, file_path, last_row_file, eligible_path):
        self.file_path = file_path
        self.last_row_file = last_row_file
        self.last_row_number = self.load_last_row_number()
        self.driver = self.setup_driver()
        self.data = {}
        self.eligible_path = eligible_path
        self.max_row = 0

    def setup_driver(self):
        chrome_options = Options()
        chrome_options.page_load_strategy = "eager"
        driver = webdriver.Chrome(
            service=ChromeService(ChromeDriverManager().install()),
            options=chrome_options,
        )
        return driver

    def load_last_row_number(self):
        try:
            with open(self.last_row_file, "r") as file:
                last_row_number = int(file.read())
            return last_row_number
        except FileNotFoundError:
            return 0

    def get_next_row_data(self):
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb.active

        # Get the company_name form the sheet
        company_name = sheet.cell(row=self.last_row_number + 2, column=1).value
        address = sheet.cell(row=self.last_row_number + 2, column=5).value
        city = sheet.cell(row=self.last_row_number + 2, column=6).value
        state = sheet.cell(row=self.last_row_number + 2, column=7).value
        return company_name, address, city, state

    def save_last_row_number(self, last_row_number):
        with open(self.last_row_file, "w") as file:
            file.write(str(last_row_number))

    def save_value(self, agent_name, agent_address):
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb.active
        sheet.cell(row=self.last_row_number + 2, column=10).value = agent_name
        sheet.cell(row=self.last_row_number + 2, column=11).value = agent_address
        wb.save(self.file_path)
        print(f"{Fore.LIGHTWHITE_EX}Details Are Saved In: {self.last_row_number + 2}")
        if name_validations.__checkValid__(self, agent_name):
            print(Fore.LIGHTRED_EX + "Not Eligible For Details")
        else:
            print(Fore.LIGHTBLUE_EX + "Yes Eligible For Details")
            self.eligibles(agent_name, agent_address)

        wb.save(self.file_path)

    def eligibles(self, agent_name, agent_address):
        wb = openpyxl.load_workbook(self.eligible_path)
        sheet = wb.active
        address, city, state, *keynames = agent_address.split(",")
        address, city, state = address.strip(), city.strip(), state.strip()
        sheet.cell(row=self.max_row + 2, column=1).value = agent_name
        sheet.cell(row=self.max_row + 2, column=5).value = address
        sheet.cell(row=self.max_row + 2, column=6).value = city
        sheet.cell(row=self.max_row + 2, column=7).value = state
        sheet.cell(row=self.max_row + 2, column=12).value = self.last_row_number + 2

        self.max_row += 1
        wb.save(self.eligible_path)

    def search_company(self):
        self.enteringDetails = True
        while self.enteringDetails:
            company_name, address, city, state = self.get_next_row_data()
            if not company_name:
                print(Fore.LIGHTBLUE_EX + "No More Rows To Process")
                self.driver.quit()
            if address_validations.__checkValid__(self, address, city, state):
                print(Fore.YELLOW + "Address Matched.")
                # save the agent details
                try:
                    with open("AgentDetails.json", "r") as infile:
                        self.data = json.load(infile)
                        agent_name = self.data["AgentDetails"]["agent_name"]
                        agent_address = self.data["AgentDetails"]["agent_address"]
                        self.save_value(agent_name, agent_address)
                        print(
                            Fore.LIGHTYELLOW_EX
                            + "\n============================================================"
                        )
                except json.JSONDecodeError:
                    print(Fore.LIGHTRED_EX + "Details are empty")
                    print(
                        Fore.LIGHTYELLOW_EX
                        + "\n============================================================"
                    )
                except FileNotFoundError:
                    print(Fore.LIGHTRED_EX + "Details are empty")
                    print(
                        Fore.LIGHTYELLOW_EX
                        + "\n============================================================"
                    )
                except:
                    print(Fore.LIGHTRED_EX + "Details are empty")
                    print(
                        Fore.LIGHTYELLOW_EX
                        + "\n============================================================"
                    )

            else:
                self.driver.implicitly_wait(10)
                # wait for the search bar to appear
                addres_bar = WebDriverWait(self.driver, 40).until(
                    EC.presence_of_element_located((By.NAME, "q"))
                )
                company_name = company_name.replace("&amp;", "&").replace("amp;", "&")
                for i in company_name:
                    addres_bar.send_keys(i)
                addres_bar.send_keys(Keys.ENTER)
                self.driver.implicitly_wait(10)
                # clicking on the minimize button
                try:
                    WebDriverWait(self.driver, 1).until(
                        EC.element_to_be_clickable(
                            (By.XPATH, "/html/body/div[2]/div[1]/div[2]/div[2]/button")
                        )
                    ).click()
                except:
                    pass
                # Locate all <li> elements
                titles = self.driver.find_elements(
                    By.CLASS_NAME, "company_search_result"
                )
                output_names = [names.text for names in titles]
                if not output_names:
                    print(
                        Back.LIGHTRED_EX
                        + Fore.BLACK
                        + "\nNO SEARCH RESULTS ARE AVAILABLE"
                    )
                    if os.path.exists("AgentDetails.json"):
                        os.remove("AgentDetails.json")
                for i, a_tag in enumerate(output_names):
                    try:
                        tag = WebDriverWait(self.driver, 5).until(
                            EC.element_to_be_clickable(
                                (By.XPATH, f"//*[contains(text(), '{a_tag}')]")
                            )
                        )

                        print(f"{Fore.GREEN}Item {i+1} <a> Text: {a_tag}")
                        tag.click()

                        # Extract agent information
                        try:
                            agent_name = WebDriverWait(self.driver, 3).until(
                                EC.presence_of_element_located(
                                    (By.CLASS_NAME, "agent_name")
                                )
                            )
                            print(Fore.LIGHTYELLOW_EX + "Agent Name is captured")
                            agent_address = WebDriverWait(self.driver, 3).until(
                                EC.presence_of_element_located(
                                    (By.CLASS_NAME, "agent_address")
                                )
                            )
                            print(Fore.LIGHTYELLOW_EX + "Agent Address is captured")
                            with open("AgentDetails.json", "w") as outfile:
                                json.dump(
                                    {
                                        "AgentDetails": {
                                            "agent_name": agent_name.text,
                                            "agent_address": agent_address.text,
                                        }
                                    },
                                    outfile,
                                )
                            self.save_value(agent_name.text, agent_address.text)
                            print(Fore.GREEN + "Values are saved!!!")
                            print(
                                Fore.LIGHTYELLOW_EX
                                + "\n============================================================"
                            )
                            break  # Break the loop after finding the agent details

                        except TimeoutException:
                            print(
                                Fore.RED
                                + "Agent details not found. Moving to the next item."
                            )

                        # Go back to the search results
                        self.driver.back()
                        self.driver.implicitly_wait(3)

                    except (NoSuchElementException, StaleElementReferenceException):
                        print(
                            f"Could not interact with item {i+1}. Moving to the next item."
                        )
                        continue
            self.last_row_number += 1
            self.save_last_row_number(self.last_row_number)
            self.driver.get(
                "https://www.opencorporates.com/"
            )  # Go back to search results

    def run(self):
        self.driver.get("https://www.opencorporates.com/")
        self.search_company()


file_path = "companies.xlsx"  # file path for search results
last_row_file = "last_row_number.txt"  # file path for last row for each iteration
eligible_path = "Eligibles/Eligibles.xlsx"  # file path for eligibles sheet
if __name__ == "__main__":
    iterations = CompanySearches(file_path, last_row_file, eligible_path)
    iterations.run()
